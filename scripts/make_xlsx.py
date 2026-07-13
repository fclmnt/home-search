#!/usr/bin/env python3
"""Régénère annonces.xlsx à partir de annonces.csv (consultable dans Numbers)."""
import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(BASE, "annonces.csv")
XLSX_PATH = os.path.join(BASE, "annonces.xlsx")

COL_WIDTHS = {
    "date_ajout": 12, "statut": 10, "titre": 40, "quartier": 22,
    "adresse": 30, "prix": 9, "superficie_pi2": 13, "chambres": 10,
    "balcon": 8, "station_metro": 18, "ligne_metro": 11,
    "minutes_a_pied": 14, "site": 14, "lien": 45, "score": 7, "notes": 50,
}

def main():
    if not os.path.exists(CSV_PATH):
        print(f"Fichier introuvable : {CSV_PATH}", file=sys.stderr)
        sys.exit(1)

    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        print("CSV vide", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Annonces"

    header = rows[0]
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    new_fill = PatternFill("solid", fgColor="E2F0D9")

    for c, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(name, 15)

    lien_idx = header.index("lien") + 1 if "lien" in header else None
    statut_idx = header.index("statut") if "statut" in header else None

    for r, row in enumerate(rows[1:], 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(header[c - 1] in ("titre", "notes")))
            if lien_idx and c == lien_idx and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
        if statut_idx is not None and len(row) > statut_idx and row[statut_idx].upper() == "NOUVEAU":
            for c in range(1, len(header) + 1):
                ws.cell(row=r, column=c).fill = new_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_PATH)
    print(f"OK : {XLSX_PATH} ({len(rows) - 1} annonces)")

if __name__ == "__main__":
    main()
